
import streamlit as st
import pandas as pd
import numpy as np
import io
import math
import datetime as dt
from typing import Dict, Tuple, List, Any
import openpyxl  # noqa: F401  # Ensure Excel engine is available
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import plotly.graph_objects as go

st.set_page_config(
    page_title="経営計画策定（単年）｜Streamlit",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded"
)

THEME_COLORS: Dict[str, str] = {
    "background": "#F7F9FB",
    "surface": "#FFFFFF",
    "surface_alt": "#E8F1FA",
    "primary": "#1F4E79",
    "primary_light": "#4F83B3",
    "accent": "#F2C57C",
    "positive": "#70A9A1",
    "positive_strong": "#2B7A78",
    "negative": "#F28F8F",
    "neutral": "#C2D3E5",
    "text": "#203040",
    "text_subtle": "#596B7A",
}

CUSTOM_STYLE = f"""
<style>
:root {{
    --base-bg: {THEME_COLORS["background"]};
    --surface: {THEME_COLORS["surface"]};
    --surface-alt: {THEME_COLORS["surface_alt"]};
    --primary: {THEME_COLORS["primary"]};
    --primary-light: {THEME_COLORS["primary_light"]};
    --accent: {THEME_COLORS["accent"]};
    --positive: {THEME_COLORS["positive"]};
    --positive-strong: {THEME_COLORS["positive_strong"]};
    --negative: {THEME_COLORS["negative"]};
    --neutral: {THEME_COLORS["neutral"]};
    --text-color: {THEME_COLORS["text"]};
    --text-subtle: {THEME_COLORS["text_subtle"]};
}}

html, body, [data-testid="stAppViewContainer"] {{
    background-color: var(--base-bg);
    color: var(--text-color);
    font-family: "Noto Sans JP", "Hiragino Sans", "Yu Gothic", sans-serif;
}}

[data-testid="stSidebar"] {{
    background: linear-gradient(180deg, var(--primary) 0%, var(--primary-light) 100%);
    color: #F7FAFC;
}}

[data-testid="stSidebar"] * {{
    color: #F7FAFC !important;
}}

.stTabs [role="tablist"] {{
    gap: 0.4rem;
    border-bottom: 1px solid var(--neutral);
}}

.stTabs [role="tab"] {{
    font-weight: 600;
    padding: 0.85rem 1.4rem;
    border-radius: 14px 14px 0 0;
    background-color: transparent;
    color: var(--text-subtle);
}}

.stTabs [role="tab"][aria-selected="true"] {{
    background-color: var(--surface);
    color: var(--primary);
    box-shadow: 0 -2px 20px rgba(31, 78, 121, 0.08);
    border-bottom: 3px solid var(--accent);
}}

div[data-testid="stMetric"] {{
    background: linear-gradient(135deg, var(--surface) 0%, var(--surface-alt) 100%);
    border-radius: 18px;
    padding: 1.15rem 1.3rem;
    box-shadow: 0 14px 28px rgba(31, 78, 121, 0.08);
    backdrop-filter: blur(6px);
}}

div[data-testid="stMetric"] [data-testid="stMetricLabel"] {{
    font-size: 0.92rem;
    color: var(--text-subtle);
}}

div[data-testid="stMetric"] [data-testid="stMetricValue"] {{
    color: var(--primary);
    font-weight: 700;
}}

div[data-testid="stMetric"] [data-testid="stMetricDelta"] {{
    color: var(--accent) !important;
}}

div[data-testid="stDataFrame"] {{
    background-color: var(--surface);
    border-radius: 18px;
    padding: 0.6rem 0.8rem 0.9rem 0.8rem;
    box-shadow: 0 12px 26px rgba(31, 78, 121, 0.06);
}}

button[kind="primary"] {{
    background-color: var(--primary);
    border-radius: 999px;
    border: none;
    box-shadow: 0 10px 20px rgba(31, 78, 121, 0.15);
}}

button[kind="primary"]:hover {{
    background-color: var(--primary-light);
}}

.hero-card {{
    background: linear-gradient(135deg, rgba(93, 169, 233, 0.92) 0%, rgba(112, 169, 161, 0.92) 100%);
    color: #ffffff;
    padding: 2.2rem 2.8rem;
    border-radius: 26px;
    box-shadow: 0 24px 48px rgba(22, 60, 90, 0.25);
    margin-bottom: 1.5rem;
}}

.hero-card h1 {{
    margin: 0 0 0.6rem 0;
    font-size: 2.35rem;
    font-weight: 700;
}}

.hero-card p {{
    margin: 0;
    font-size: 1.08rem;
    opacity: 0.92;
}}

.insight-card {{
    background-color: var(--surface);
    border-radius: 18px;
    padding: 1.1rem 1.3rem;
    box-shadow: 0 12px 24px rgba(31, 78, 121, 0.08);
    border-left: 6px solid var(--primary-light);
    margin-bottom: 1rem;
}}

.insight-card.positive {{
    border-left-color: var(--positive);
}}

.insight-card.warning {{
    border-left-color: var(--accent);
}}

.insight-card.alert {{
    border-left-color: var(--negative);
}}

.insight-card h4 {{
    margin: 0 0 0.4rem 0;
    font-size: 1.05rem;
    color: var(--primary);
}}

.insight-card p {{
    margin: 0;
    font-size: 0.95rem;
    color: var(--text-subtle);
    line-height: 1.55;
}}

div[data-testid="stDataFrame"] table {{
    border-spacing: 0;
    color: var(--text-color);
}}

div[data-testid="stDataFrame"] table thead th {{
    background: rgba(31, 78, 121, 0.08);
    color: var(--primary);
    font-weight: 600;
    border-bottom: 1px solid rgba(31, 78, 121, 0.18);
}}

div[data-testid="stDataFrame"] table tbody tr:nth-child(even) {{
    background: rgba(31, 78, 121, 0.04);
}}

div[data-testid="stDataFrame"] table tbody tr:hover {{
    background: rgba(242, 197, 124, 0.16);
    transition: background 0.2s ease;
}}

.cost-pill {{
    background: linear-gradient(135deg, rgba(79, 131, 179, 0.16) 0%, rgba(31, 78, 121, 0.08) 100%);
    border-radius: 18px;
    padding: 1rem 1.2rem;
    margin-bottom: 0.8rem;
    display: flex;
    flex-direction: column;
    gap: 0.35rem;
    border: 1px solid rgba(31, 78, 121, 0.12);
}}

.cost-pill.positive {{
    background: linear-gradient(135deg, rgba(112, 169, 161, 0.18) 0%, rgba(112, 169, 161, 0.08) 100%);
    border-color: rgba(112, 169, 161, 0.3);
}}

.cost-pill.accent {{
    background: linear-gradient(135deg, rgba(242, 197, 124, 0.18) 0%, rgba(242, 197, 124, 0.1) 100%);
    border-color: rgba(242, 197, 124, 0.32);
}}

.cost-pill strong {{
    font-size: 1.05rem;
    color: var(--primary);
}}

.cost-pill span {{
    font-size: 1.15rem;
    font-weight: 700;
    color: var(--text-color);
}}

.cost-pill small {{
    font-size: 0.85rem;
    color: var(--text-subtle);
}}

.glossary-card {{
    background: linear-gradient(135deg, rgba(255, 255, 255, 0.95) 0%, rgba(232, 241, 250, 0.95) 100%);
    border-radius: 18px;
    padding: 1.2rem 1.4rem;
    box-shadow: 0 10px 26px rgba(31, 78, 121, 0.06);
    border: 1px solid rgba(31, 78, 121, 0.08);
    margin-top: 1.2rem;
}}

.glossary-card h4 {{
    margin: 0 0 0.6rem 0;
    color: var(--primary);
}}

.glossary-card ul {{
    margin: 0;
    padding-left: 1.1rem;
    display: grid;
    gap: 0.45rem;
}}

.glossary-card li {{
    list-style: none;
    background: rgba(31, 78, 121, 0.05);
    border-radius: 12px;
    padding: 0.75rem 0.9rem;
    border: 1px solid rgba(31, 78, 121, 0.08);
}}

.glossary-card li strong {{
    display: block;
    font-weight: 600;
    color: var(--primary);
    margin-bottom: 0.15rem;
}}

.glossary-card li span {{
    font-size: 0.92rem;
    color: var(--text-subtle);
}}

.anomaly-table caption {{
    caption-side: top;
    font-weight: 600;
    color: var(--primary);
}}

.ai-badge {{
    display: inline-flex;
    align-items: center;
    gap: 0.4rem;
    background-color: rgba(112, 169, 161, 0.16);
    color: var(--positive-strong);
    border-radius: 999px;
    padding: 0.35rem 0.9rem;
    font-weight: 600;
    letter-spacing: 0.02em;
}}
</style>
"""

st.markdown(CUSTOM_STYLE, unsafe_allow_html=True)

st.markdown(
    """
    <div class="hero-card">
        <h1>McKinsey Inspired 経営計画ダッシュボード</h1>
        <p>直感的な操作とAI分析で、戦略から実行までを素早くデザインします。グラフ・KPI・シナリオを洗練されたUIで俯瞰し、最適な意思決定をサポートします。</p>
    </div>
    """,
    unsafe_allow_html=True,
)

DEFAULTS = {
    "sales": 1000000000,
    "fte": 20.0,
    "cogs_mat_rate": 0.25,
    "cogs_lbr_rate": 0.06,
    "cogs_out_src_rate": 0.1,
    "cogs_out_con_rate": 0.04,
    "cogs_oth_rate": 0.0,
    "opex_h_rate": 0.17,
    "opex_k_rate": 0.468,
    "opex_dep_rate": 0.006,
    "noi_misc_rate": 0.0001,
    "noi_grant_rate": 0.0,
    "noi_oth_rate": 0.0,
    "noe_int_rate": 0.0074,
    "noe_oth_rate": 0.0,
    "unit": "百万円",
    "fiscal_year": 2025
}

PLOT_STYLE_DEFAULT: Dict[str, Any] = {
    "figure_bg": THEME_COLORS["surface"],
    "axes_bg": THEME_COLORS["surface"],
    "grid": True,
    "grid_color": "#D4DEE9",
    "pos_color": THEME_COLORS["positive"],
    "neg_color": THEME_COLORS["negative"],
    "node_size": 11,
    "font_color": THEME_COLORS["text"],
    "font_size": 11,
    "alpha": 0.88,
}

# --- 営業外の既定値（必要に応じてサイドバー入力にしても良い） ---
NONOP_DEFAULT = dict(
    noi_misc=0.0,
    noi_grant=0.0,
    noe_int=0.0,
    noe_oth=0.0,
)

ITEMS = [
    ("REV", "売上高", "売上"),
    ("COGS_MAT", "外部仕入｜材料費", "外部仕入"),
    ("COGS_LBR", "外部仕入｜労務費(外部)", "外部仕入"),
    ("COGS_OUT_SRC", "外部仕入｜外注費(専属)", "外部仕入"),
    ("COGS_OUT_CON", "外部仕入｜外注費(委託)", "外部仕入"),
    ("COGS_OTH", "外部仕入｜その他諸経費", "外部仕入"),
    ("COGS_TTL", "外部仕入｜計", "外部仕入"),
    ("GROSS", "粗利(加工高)", "粗利"),
    ("OPEX_H", "内部費用｜人件費", "内部費用"),
    ("OPEX_K", "内部費用｜経費", "内部費用"),
    ("OPEX_DEP", "内部費用｜減価償却費", "内部費用"),
    ("OPEX_TTL", "内部費用｜計", "内部費用"),
    ("OP", "営業利益", "損益"),
    ("NOI_MISC", "営業外収益｜雑収入", "営業外"),
    ("NOI_GRANT", "営業外収益｜補助金/給付金", "営業外"),
    ("NOI_OTH", "営業外収益｜その他", "営業外"),
    ("NOE_INT", "営業外費用｜支払利息", "営業外"),
    ("NOE_OTH", "営業外費用｜雑損", "営業外"),
    ("ORD", "経常利益", "損益"),
    ("BE_SALES", "損益分岐点売上高", "KPI"),
    ("PC_SALES", "一人当たり売上", "KPI"),
    ("PC_GROSS", "一人当たり粗利", "KPI"),
    ("PC_ORD", "一人当たり経常利益", "KPI"),
    ("LDR", "労働分配率", "KPI")
]

# Mapping from item code to label for quick lookup
ITEM_LABELS = {code: label for code, label, _ in ITEMS}

PLAIN_LANGUAGE = {
    "REV": "お客様から入る売上全体",  # Revenue
    "COGS_MAT": "主原料や仕入にかかるコスト",
    "COGS_LBR": "外部スタッフや職人さんへの人件費",
    "COGS_OUT_SRC": "専属パートナーへの外注費",
    "COGS_OUT_CON": "必要時だけ依頼するスポット外注費",
    "COGS_OTH": "物流・包材などその他の変動費",
    "COGS_TTL": "外部仕入コストの合計",
    "GROSS": "売上から原価を引いた稼ぐ力（CT）",
    "OPEX_H": "自社の人件費（給与・賞与など）",
    "OPEX_K": "オフィス費や販売促進費などの経費",
    "OPEX_DEP": "設備投資を分割計上した費用",
    "OPEX_TTL": "内部費用の合計",
    "OP": "本業だけで稼いだ利益",
    "NOI_MISC": "本業以外の雑収入",
    "NOI_GRANT": "補助金・給付金などの臨時収入",
    "NOI_OTH": "その他の営業外収益",
    "NOE_INT": "借入金などの利息支払",
    "NOE_OTH": "その他の営業外費用",
    "ORD": "金融費用も含めた最終的な利益",
    "BE_SALES": "利益がプラスに転じる売上ライン",
    "PC_SALES": "1人あたりの売上高",
    "PC_GROSS": "1人あたりの粗利（CT）",
    "PC_ORD": "1人あたりの経常利益",
    "LDR": "粗利のうち人件費に充てている割合",
}

COST_PILL_ITEMS = [
    ("COGS_MAT", "材料費", "製品づくりに必要な仕入原価", ""),
    ("COGS_LBR", "外部人件費", "外部メンバーへの人件費", ""),
    ("COGS_OUT_SRC", "協力会社費（専属）", "固定契約パートナーへの支払い", ""),
    ("COGS_OUT_CON", "協力会社費（スポット）", "必要なタイミングだけの外注費", ""),
    ("COGS_OTH", "その他原価", "物流費や包材などの付随コスト", ""),
    ("COGS_TTL", "標準原価 合計", "外部仕入コストの総額", "accent"),
    ("GROSS", "粗利（CT）", "原価を差し引いた稼ぐ力", "positive"),
]

GLOSSARY_ITEMS = [
    {"term": "CT（粗利）", "description": "Contribution Marginの略。売上から変動費を引いた稼ぐ力を指します。"},
    {"term": "標準原価", "description": "製品・サービスを提供するために想定される平均的な原価のこと。材料費や外注費を含みます。"},
    {"term": "営業利益", "description": "本業でどれだけ利益が残ったかを示します。粗利から人件費や経費を差し引いた金額です。"},
    {"term": "経常利益", "description": "営業利益に利息収支などの営業外項目を加減した企業全体の稼ぐ力です。"},
    {"term": "損益分岐点売上高", "description": "利益がゼロになる境目の売上高。ここを超えると利益が積み上がります。"},
    {"term": "労働分配率", "description": "粗利のうち、どれだけを人件費として従業員へ配分しているかを示す指標です。"},
]

# --- MCKINSEY TORNADO
def _set_jp_font() -> None:
    """日本語フォントを自動設定（環境に応じて存在チェック）"""
    for f in ["Yu Gothic", "Meiryo", "Hiragino Sans", "Noto Sans CJK JP", "IPAexGothic"]:
        try:
            mpl.font_manager.findfont(f, fallback_to_default=False)
            mpl.rcParams["font.family"] = f
            break
        except Exception:
            continue
    mpl.rcParams["axes.unicode_minus"] = False

def render_tornado_mckinsey(
    changes: List[Tuple[str, float]],
    title: str,
    unit_label: str,
    style: Dict[str, Any] | None = None,
) -> None:
    """マッキンゼー風トルネード図を描画しPNGダウンロードボタンを表示"""
    style = style or {}
    fig_bg = style.get("figure_bg", PLOT_STYLE_DEFAULT["figure_bg"])
    axes_bg = style.get("axes_bg", PLOT_STYLE_DEFAULT["axes_bg"])
    grid_on = style.get("grid", PLOT_STYLE_DEFAULT["grid"])
    grid_color = style.get("grid_color", PLOT_STYLE_DEFAULT["grid_color"])
    pos_color = style.get("pos_color", PLOT_STYLE_DEFAULT["pos_color"])
    neg_color = style.get("neg_color", PLOT_STYLE_DEFAULT["neg_color"])
    node_size = style.get("node_size", PLOT_STYLE_DEFAULT["node_size"])
    font_color = style.get("font_color", PLOT_STYLE_DEFAULT["font_color"])
    font_size = style.get("font_size", PLOT_STYLE_DEFAULT["font_size"])
    alpha = style.get("alpha", PLOT_STYLE_DEFAULT["alpha"])
    if not changes:
        st.warning("表示するデータがありません。")
        return
    changes_sorted = sorted(changes, key=lambda x: abs(x[1]), reverse=True)
    labels = [k for k, _ in changes_sorted]
    values = [v for _, v in changes_sorted]
    max_abs = max(abs(v) for v in values)
    if not math.isfinite(max_abs) or max_abs == 0:
        st.warning("有効なデータがありません。")
        return
    lim = max_abs * 1.1
    fig, ax = plt.subplots(figsize=(6, 0.45 * len(values) + 1))
    fig.patch.set_facecolor(fig_bg)
    ax.set_facecolor(axes_bg)
    y = np.arange(len(values))
    colors = [pos_color if v >= 0 else neg_color for v in values]
    bars = ax.barh(y, values, color=colors, alpha=alpha)
    ax.set_yticks(y, labels)
    ax.set_xlim(-lim, lim)
    ax.axvline(0, color=grid_color, linewidth=0.8)
    if grid_on:
        ax.grid(color=grid_color, axis="x", linewidth=0.5, linestyle="--")
    else:
        ax.grid(False)
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)
    for spine in ("left", "bottom"):
        ax.spines[spine].set_color(grid_color)
        ax.spines[spine].set_linewidth(0.5)
    ax.xaxis.set_major_formatter(FuncFormatter(lambda x, _: f"¥{x:,.0f}"))
    ax.tick_params(axis="x", colors=font_color, labelsize=font_size)
    ax.tick_params(axis="y", colors=font_color, labelsize=font_size)
    ellipsis = False
    for bar, v in zip(bars, values):
        txt = f"{'+' if v >= 0 else '-'}¥{abs(v):,}"
        if abs(v) < lim * 0.05:
            txt = "..."
            ellipsis = True
        ax.text(
            v + (lim * 0.01 if v >= 0 else -lim * 0.01),
            bar.get_y() + bar.get_height() / 2,
            txt,
            ha="left" if v >= 0 else "right",
            va="center",
            clip_on=False,
            color=font_color,
            fontsize=node_size,
        )
    ax.set_title(title, color=font_color, fontsize=font_size + 2)
    fig.tight_layout()
    fig.text(0.5, -0.02, "注：右=利益増、左=利益減", ha="center", fontsize=font_size - 1, color=font_color)
    st.pyplot(fig, use_container_width=True)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=200, bbox_inches="tight")
    st.download_button(
        "📥 感応度グラフ（PNG）",
        data=buf.getvalue(),
        file_name="tornado.png",
        mime="image/png",
    )
    if ellipsis:
        st.caption("※ 一部の値は省略記号で表示しています。下表で詳細を確認ください。")


def build_sensitivity_view_options(
    parent: st.delta_generator.DeltaGenerator | None = None,
    *,
    key_prefix: str = "sensitivity",
    defaults: Dict[str, Any] | None = None,
    show_header: bool = True,
) -> Dict[str, Any]:
    """感応度グラフの各種コントロール（配置可能なように柔軟化）。"""

    ctx = parent if parent is not None else st
    defaults = defaults or {}
    options = ["トルネード（±差分）", "ウォーターフォール（寄与累積）"]
    default_viz = defaults.get("viz", options[0])
    viz_index = options.index(default_viz) if default_viz in options else 0

    if show_header:
        ctx.subheader("📉 感応度分析｜表示設定")

    c1, c2, c3, c4 = ctx.columns([2, 1.1, 1.1, 1.0])
    with c1:
        viz = st.radio(
            "可視化タイプ",
            options,
            horizontal=True,
            index=viz_index,
            key=f"{key_prefix}_viz",
        )
    with c2:
        top_n = st.slider(
            "表示項目数 (Top-N)",
            3,
            12,
            int(defaults.get("top_n", 6)),
            1,
            key=f"{key_prefix}_topn",
        )
    with c3:
        height_px = st.slider(
            "グラフ高さ (px)",
            200,
            900,
            int(defaults.get("height_px", 360)),
            20,
            key=f"{key_prefix}_height",
        )
    with c4:
        compact = st.checkbox(
            "コンパクト表示（小さな文字）",
            value=bool(defaults.get("compact", True)),
            key=f"{key_prefix}_compact",
        )

    step = ctx.slider(
        "感応度ステップ（±）",
        0.01,
        0.20,
        float(defaults.get("step", 0.10)),
        0.01,
        key=f"{key_prefix}_step",
    )
    show_values = ctx.checkbox(
        "値ラベルを表示",
        value=bool(defaults.get("show_values", True)),
        key=f"{key_prefix}_showvalues",
    )
    return dict(
        viz=viz,
        top_n=top_n,
        height_px=height_px,
        compact=compact,
        step=step,
        show_values=show_values,
    )


def _sensitivity_items(plan: dict, step: float):
    """各変数の±stepによる経常利益差分（ceteris paribus）。"""
    keys = [
        ("sales", "売上高", "amount"),
        ("gp_rate", "粗利率", "rate"),
        ("opex_h", "人件費", "amount"),
        ("opex_fixed", "販管費（固定費）", "amount"),
        ("opex_dep", "減価償却", "amount"),
        ("opex_oth", "その他費用", "amount"),
    ]
    base_ord = compute_plan(plan)["ord"]
    items = []
    for k, label, kind in keys:
        p_low = plan.copy()
        p_high = plan.copy()
        if kind == "rate":
            p_low[k] = max(0.0, plan[k] - step)
            p_high[k] = min(1.0, plan[k] + step)
        else:
            p_low[k] = max(0.0, plan[k] * (1 - step))
            p_high[k] = plan[k] * (1 + step)

        low_ord = compute_plan(p_low)["ord"]
        high_ord = compute_plan(p_high)["ord"]
        delta_low = low_ord - base_ord
        delta_high = high_ord - base_ord
        span = abs(delta_low) + abs(delta_high)
        items.append(dict(key=k, label=label,
                          delta_low=delta_low, delta_high=delta_high, span=span))
    items.sort(key=lambda x: x["span"], reverse=True)
    return items


def render_tornado_compact(plan: dict, step: float, top_n: int, height_px: int,
                           compact: bool, show_values: bool):
    """俯瞰性を高めたトルネード図（Top-N・高さ・フォント調整）"""
    items = _sensitivity_items(plan, step)[:top_n]
    labels = [x["label"] for x in items]
    lows = [x["delta_low"] for x in items]
    highs = [x["delta_high"] for x in items]

    fig_h_in = max(height_px / 96.0, 2 / 3)
    fig, ax = plt.subplots(figsize=(7, fig_h_in))
    for i, (lo, hi) in enumerate(zip(lows, highs)):
        hi_color = THEME_COLORS["positive"] if hi >= 0 else THEME_COLORS["negative"]
        lo_color = THEME_COLORS["negative"] if lo <= 0 else THEME_COLORS["positive"]
        ax.barh(i, hi, color=hi_color, alpha=0.85)
        ax.barh(i, lo, color=lo_color, alpha=0.45)

    ax.set_yticks(range(len(labels)))
    ax.set_yticklabels(labels, fontsize=(9 if compact else 11))
    ax.xaxis.set_major_formatter(FuncFormatter(lambda x, _: f"¥{x:,.0f}"))
    ax.axvline(0, color=THEME_COLORS["neutral"], linewidth=1.0, linestyle="--")
    ax.set_xlabel("経常利益への寄与（差分）", fontsize=(9 if compact else 11))

    if show_values:
        offset = max(1.0, max(abs(v) for v in lows + highs) * 0.02)
        for i, (lo, hi) in enumerate(zip(lows, highs)):
            ax.text(hi + (offset if hi >= 0 else -offset),
                    i, format_money(hi), va="center", ha="left" if hi >= 0 else "right", fontsize=(8 if compact else 10), color=THEME_COLORS["text"])
            ax.text(lo + (offset if lo >= 0 else -offset),
                    i, format_money(lo), va="center", ha="left" if lo >= 0 else "right", fontsize=(8 if compact else 10), color=THEME_COLORS["text_subtle"])

    fig.tight_layout()
    st.pyplot(fig, use_container_width=True)


def render_sensitivity_waterfall(plan: dict, step: float, top_n: int, height_px: int,
                                 compact: bool, show_values: bool):
    """
    感応度の「寄与累積」をウォーターフォールで表示。
    ・各変数を +step 側に単独シフトした場合の寄与を絶対値降順に並べ、
      ベースORDから順に累積表示（相互作用は考慮しない近似）。
    """
    base_ord = compute_plan(plan)["ord"]
    items = _sensitivity_items(plan, step)[:top_n]
    contribs = [(x["label"], x["delta_high"]) for x in items]
    labels = ["ベースORD"] + [lbl for lbl, _ in contribs] + ["概算ORD（+step適用）"]
    vals = [base_ord] + [v for _, v in contribs] + [0.0]
    cum = [vals[0]]
    for v in vals[1:-1]:
        cum.append(cum[-1] + v)
    final = cum[-1]
    vals[-1] = final - base_ord

    fig_h_in = max(height_px / 96.0, 2 / 3)
    fig, ax = plt.subplots(figsize=(7, fig_h_in))
    colors = []
    for i, v in enumerate(vals):
        if i == 0 or i == len(vals) - 1:
            colors.append(THEME_COLORS["primary"])
        else:
            colors.append(THEME_COLORS["positive"] if v >= 0 else THEME_COLORS["negative"])

    ax.bar(range(len(vals)), vals, color=colors, alpha=0.88)
    ax.axhline(0, color=THEME_COLORS["neutral"], linewidth=1.0, linestyle="--")
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=(8 if compact else 10))
    ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f"¥{x:,.0f}"))
    ax.set_ylabel("寄与（累積）", fontsize=(9 if compact else 11))

    if show_values:
        ref = max(1.0, max(abs(v) for v in vals))
        for i, v in enumerate(vals):
            ax.text(i, v + (0.02 * ref if v >= 0 else -0.02 * ref),
                    format_money(v), ha="center",
                    va="bottom" if v >= 0 else "top",
                    fontsize=(8 if compact else 10))

    fig.tight_layout()
    st.pyplot(fig, use_container_width=True)


def render_sensitivity_view(plan: dict):
    """感応度分析ビューの統括（俯瞰性改善＋ウォーターフォール追加）"""
    zoom_mode = st.toggle(
        "🔍 グラフ拡大モードで操作する",
        value=st.session_state.get("sensitivity_zoom_mode", False),
        key="sensitivity_zoom_mode",
        help="横幅いっぱいにグラフを表示しつつ、右側のコントロールでリアルタイムに調整できます。",
    )

    if zoom_mode:
        chart_col, ctrl_col = st.columns([3.2, 1.8])
        with ctrl_col:
            st.markdown("<span class='ai-badge'>Zoomコントロール</span>", unsafe_allow_html=True)
            opt = build_sensitivity_view_options(
                parent=ctrl_col,
                key_prefix="sensitivity",
                defaults=st.session_state.get("sensitivity_current", {}),
                show_header=False,
            )
            st.caption("設定は自動で保存され、通常表示に戻っても引き継がれます。")
        target_container = chart_col
    else:
        opt = build_sensitivity_view_options(key_prefix="sensitivity")
        target_container = st.container()

    st.session_state["sensitivity_current"] = opt

    with target_container:
        if zoom_mode:
            st.markdown("#### 🔎 拡大ビュー（ライブ更新）")
        if opt["viz"].startswith("トルネード"):
            render_tornado_compact(
                plan,
                opt["step"],
                opt["top_n"],
                opt["height_px"],
                opt["compact"],
                opt["show_values"],
            )
        else:
            render_sensitivity_waterfall(
                plan,
                opt["step"],
                opt["top_n"],
                opt["height_px"],
                opt["compact"],
                opt["show_values"],
            )

# --- EXCEL JP LOCALE
def apply_japanese_styles(wb) -> None:
    """ヘッダ太字・中央揃え、列幅自動調整、1行目固定"""
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(28, max(10, max_len + 2))

def format_money_and_percent(ws, money_cols: List[int], percent_cols: List[int]) -> None:
    """通貨および百分率の書式を適用"""
    money_fmt = "\"¥\"#,##0;[Red]-\"¥\"#,##0"
    for c in money_cols:
        col_letter = get_column_letter(c)
        for cell in ws[col_letter][1:]:
            cell.number_format = money_fmt
    for c in percent_cols:
        col_letter = get_column_letter(c)
        for cell in ws[col_letter][1:]:
            cell.number_format = "0%"

def millions(x):
    return x / 1_000_000

def thousands(x):
    return x / 1_000

def format_money(x, unit="百万円"):
    if x is None or (isinstance(x, float) and (np.isnan(x) or np.isinf(x))):
        return "—"
    if unit == "百万円":
        return f"{millions(x):,.0f}"
    elif unit == "千円":
        return f"{thousands(x):,.0f}"
    else:
        return f"{x:,.0f}"


def summarize_plan_metrics(amounts: Dict[str, float]) -> Dict[str, float]:
    """計画値から主要指標（率・水準）を算出。"""

    sales = float(amounts.get("REV", 0.0))
    gross = float(amounts.get("GROSS", 0.0))
    op = float(amounts.get("OP", 0.0))
    ord_val = float(amounts.get("ORD", 0.0))
    opex = float(amounts.get("OPEX_TTL", 0.0))
    cogs = float(amounts.get("COGS_TTL", sales - gross))

    def safe_ratio(num: float, den: float) -> float:
        return float(num / den) if den not in (0, None) else float("nan")

    metrics = {
        "sales": sales,
        "gross": gross,
        "op": op,
        "ord": ord_val,
        "gross_margin": safe_ratio(gross, sales),
        "op_margin": safe_ratio(op, sales),
        "ord_margin": safe_ratio(ord_val, sales),
        "cogs_ratio": safe_ratio(cogs, sales),
        "opex_ratio": safe_ratio(opex, sales),
        "labor_ratio": safe_ratio(amounts.get("OPEX_H", 0.0), gross),
        "breakeven": float(amounts.get("BE_SALES", float("nan"))),
    }
    return metrics


def format_ratio(value: float) -> str:
    if value is None or (isinstance(value, float) and (np.isnan(value) or np.isinf(value))):
        return "—"
    return f"{value * 100:.1f}%"


def generate_ai_recommendations(
    metrics: Dict[str, float],
    numeric_amounts: pd.DataFrame | None,
    numeric_kpis: pd.DataFrame | None,
    unit: str,
) -> List[Dict[str, str]]:
    """定性コメントを生成（ルールベースAIのアシスト）。"""

    insights: List[Dict[str, str]] = []
    gm = metrics.get("gross_margin")
    ord_margin = metrics.get("ord_margin")
    labor_ratio = metrics.get("labor_ratio")
    be_sales = metrics.get("breakeven")
    sales = metrics.get("sales", 0.0)

    if gm is not None and math.isfinite(gm):
        if gm < 0.25:
            insights.append({
                "title": "粗利率が低位です",
                "body": "粗利率が25%を下回っています。価格改定や高付加価値サービスの投入でマージン改善を検討しましょう。",
                "tone": "warning",
            })
        elif gm > 0.45:
            insights.append({
                "title": "粗利率はプレミアム水準",
                "body": "粗利率が45%超と高水準です。余剰利益を投資や人材育成に再配分する余地があります。",
                "tone": "positive",
            })

    if ord_margin is not None and math.isfinite(ord_margin):
        if ord_margin < 0:
            insights.append({
                "title": "経常利益が赤字レンジ",
                "body": "経常利益がマイナスです。固定費削減と利益率の高い案件へのシフトを緊急で検討してください。",
                "tone": "alert",
            })
        elif ord_margin < 0.05:
            insights.append({
                "title": "利益率の底上げが必要",
                "body": "経常利益率が5%未満です。販売単価の引き上げや高粗利商品の比率向上が改善策になります。",
                "tone": "warning",
            })
        elif ord_margin > 0.12:
            insights.append({
                "title": "利益創出力は堅調",
                "body": "経常利益率が12%超と十分な稼ぐ力があります。積極投資フェーズに移行しても耐性があります。",
                "tone": "positive",
            })

    if labor_ratio is not None and math.isfinite(labor_ratio):
        if labor_ratio > 0.65:
            insights.append({
                "title": "人件費の比率が高い",
                "body": "労働分配率が65%を超えています。生産性向上策やアウトソースの活用でコストを平準化しましょう。",
                "tone": "warning",
            })
        elif labor_ratio < 0.45:
            insights.append({
                "title": "人材投資の余地あり",
                "body": "労働分配率が45%未満です。人材強化やインセンティブ設計に投資し、組織力を底上げするチャンスです。",
                "tone": "positive",
            })

    if be_sales and sales and math.isfinite(be_sales):
        be_ratio = be_sales / sales if sales else float("nan")
        if math.isfinite(be_ratio) and be_ratio > 0.95:
            insights.append({
                "title": "損益分岐点が売上に接近",
                "body": "損益分岐点売上がほぼフル稼働の水準です。固定費の圧縮や粗利率改善で安全余裕を確保しましょう。",
                "tone": "alert",
            })
        elif math.isfinite(be_ratio) and be_ratio < 0.75:
            insights.append({
                "title": "損益分岐点に余裕あり",
                "body": "損益分岐点が売上の75%未満で、収益構造に安全余裕があります。成長投資のアクセルを踏める状態です。",
                "tone": "positive",
            })

    if numeric_amounts is not None and not numeric_amounts.empty:
        value_cols = [c for c in numeric_amounts.columns if c != "項目"]
        if len(value_cols) >= 2 and "ORD" in numeric_amounts.index:
            base_col = value_cols[0]
            base_ord = float(numeric_amounts.loc["ORD", base_col])
            best_col = None
            best_diff = 0.0
            for col in value_cols[1:]:
                diff = float(numeric_amounts.loc["ORD", col]) - base_ord
                if diff > best_diff:
                    best_diff = diff
                    best_col = col
            if best_col and best_diff > 0:
                insights.append({
                    "title": f"最有力シナリオ：{best_col}",
                    "body": f"ベース比で経常利益を{format_money(best_diff, unit)} {unit}押し上げます。主要ドライバを戦略課題に落とし込みましょう。",
                    "tone": "positive",
                })

    if not insights:
        insights.append({
            "title": "データ点検が完了しました",
            "body": "大きな懸念は検出されませんでした。引き続きシナリオ比較と感応度を活用し、計画をチューニングしてください。",
            "tone": "positive",
        })

    return insights[:5]


def detect_anomalies_in_plan(
    numeric_amounts: pd.DataFrame | None,
    numeric_kpis: pd.DataFrame | None,
    unit: str,
    metrics: Dict[str, float],
) -> pd.DataFrame:
    """異常値（高リスク・高インパクト）の候補を抽出。"""

    cols = ["カテゴリ", "対象", "値", "判定", "コメント"]
    if numeric_amounts is None or numeric_amounts.empty:
        return pd.DataFrame(columns=cols)

    value_cols = [c for c in numeric_amounts.columns if c != "項目"]
    if not value_cols:
        return pd.DataFrame(columns=cols)

    base_col = value_cols[0]
    anomalies: List[Dict[str, str]] = []

    sales = metrics.get("sales", 0.0)
    base_ord = metrics.get("ord", 0.0)
    base_op = metrics.get("op", 0.0)
    base_gross = metrics.get("gross", 0.0)
    be_sales = metrics.get("breakeven", float("nan"))

    def record(category: str, target: str, value: str, judgement: str, comment: str) -> None:
        anomalies.append({"カテゴリ": category, "対象": target, "値": value, "判定": judgement, "コメント": comment})

    if base_gross <= 0:
        record("損益", "粗利（目標）", f"{format_money(base_gross, unit)} {unit}", "🚨 粗利が不足", "売上より費用が先行しています。可変費率の再点検が必要です。")
    if base_op < 0:
        record("損益", "営業利益（目標）", f"{format_money(base_op, unit)} {unit}", "🚨 赤字リスク", "営業利益がマイナスです。固定費の削減や高粗利案件へのシフトを優先してください。")
    if base_ord < 0:
        record("損益", "経常利益（目標）", f"{format_money(base_ord, unit)} {unit}", "🚨 経常赤字", "営業外損益も含め赤字レンジです。財務・本業双方のてこ入れが求められます。")

    gm = metrics.get("gross_margin")
    if gm is not None and math.isfinite(gm) and gm < 0.2:
        record("利益率", "粗利率", format_ratio(gm), "⚠️ マージン低下", "粗利率が20%を割り込んでいます。価格戦略や原価低減を検討してください。")

    ldr_value = None
    if numeric_kpis is not None and not numeric_kpis.empty and "LDR" in numeric_kpis.index:
        ldr_value = float(numeric_kpis.loc["LDR", base_col])
        if math.isfinite(ldr_value) and ldr_value > 0.7:
            record("人件費", "労働分配率（目標）", format_ratio(ldr_value), "⚠️ 人件費過多", "人件費比率が高すぎます。工数マネジメントや外注活用でバランスを取りましょう。")

    cogs_ratio = metrics.get("cogs_ratio")
    if cogs_ratio is not None and math.isfinite(cogs_ratio) and cogs_ratio > 0.8:
        record("コスト構造", "外部仕入比率", format_ratio(cogs_ratio), "⚠️ コスト高止まり", "仕入費用が売上の80%超です。サプライヤー交渉やポートフォリオ見直しが必要です。")

    if be_sales and sales and math.isfinite(be_sales) and be_sales > sales * 0.95:
        record("安全余裕", "損益分岐点売上", f"{format_money(be_sales, unit)} {unit}", "⚠️ 余裕が僅少", "損益分岐点が現計画売上の95%超です。固定費圧縮で安全マージンを確保しましょう。")

    if numeric_amounts is not None and "ORD" in numeric_amounts.index and len(value_cols) > 1:
        base_ord_value = float(numeric_amounts.loc["ORD", base_col])
        baseline = max(abs(base_ord_value), sales * 0.02, 1_000_000.0)
        for col in value_cols[1:]:
            scn_value = float(numeric_amounts.loc["ORD", col])
            diff = scn_value - base_ord_value
            if diff <= -0.5 * baseline:
                record(
                    "シナリオ",
                    f"{col}｜経常利益",
                    f"{format_money(scn_value, unit)} {unit}",
                    "🚨 大幅悪化",
                    f"ベース比で{format_money(abs(diff), unit)} {unit}の減益です。前提条件の見直しが必要です。",
                )
            elif diff >= 0.5 * baseline:
                record(
                    "シナリオ",
                    f"{col}｜経常利益",
                    f"{format_money(scn_value, unit)} {unit}",
                    "✅ 大幅改善",
                    f"ベース比で{format_money(diff, unit)} {unit}増益です。実現可能性と投資アクションを検証しましょう。",
                )

    if numeric_kpis is not None and not numeric_kpis.empty and "LDR" in numeric_kpis.index and len(value_cols) > 1:
        for col in value_cols[1:]:
            ldr = float(numeric_kpis.loc["LDR", col])
            if math.isfinite(ldr) and ldr > 0.75:
                record(
                    "人件費",
                    f"{col}｜労働分配率",
                    format_ratio(ldr),
                    "⚠️ 人件費過重",
                    "シナリオ適用時に人件費比率が75%を超えます。追加施策での吸収が必要です。",
                )

    if not anomalies:
        return pd.DataFrame(columns=cols)

    return pd.DataFrame(anomalies, columns=cols)


def compute_plan(plan: dict) -> dict:
    """簡易計画計算。

    sales, gp_rate, 各種費用を受け取り、粗利や営業利益を算出する。
    戻り値には計算過程の主要項目を含める。
    """
    sales = float(plan.get("sales", 0.0))
    gp_rate = float(plan.get("gp_rate", 0.0))
    gross = sales * gp_rate
    opex_h = float(plan.get("opex_h", 0.0))
    opex_fixed = float(plan.get("opex_fixed", 0.0))
    opex_dep = float(plan.get("opex_dep", 0.0))
    opex_oth = float(plan.get("opex_oth", 0.0))
    op = gross - opex_h - opex_fixed - opex_dep - opex_oth
    return {
        "sales": sales,
        "gp_rate": gp_rate,
        "gross": gross,
        "opex_h": opex_h,
        "opex_fixed": opex_fixed,
        "opex_dep": opex_dep,
        "opex_oth": opex_oth,
        "op": op,
        "ord": op,
    }


def _ord_from(res: dict, nonop: dict) -> float:
    """OP と 営業外から ORD を算出"""
    noi = (nonop.get("noi_misc", 0.0) + nonop.get("noi_grant", 0.0))
    noe = (nonop.get("noe_int", 0.0) + nonop.get("noe_oth", 0.0))
    return res["op"] + noi - noe


def _plan_with(plan: dict, **overrides) -> dict:
    p = plan.copy()
    p.update(overrides)
    return p


def _line_items(res: dict, nonop: dict) -> dict:
    """行定義を一元化（REV/COGS/GROSS/OPEX/OP/営業外/ORD）"""
    rev = res["sales"]
    gross = res["gross"]
    cogs_ttl = rev - gross
    opex_ttl = res["opex_fixed"] + res["opex_h"] + res["opex_dep"] + res["opex_oth"]
    ord_v = _ord_from(res, nonop)
    return {
        "REV": rev,
        "COGS_TTL": cogs_ttl,
        "GROSS": gross,
        "OPEX_H": res["opex_h"],
        "OPEX_FIXED": res["opex_fixed"],
        "OPEX_DEP": res["opex_dep"],
        "OPEX_OTH": res["opex_oth"],
        "OPEX_TTL": opex_ttl,
        "OP": res["op"],
        "NOI_MISC": nonop.get("noi_misc", 0.0),
        "NOI_GRANT": nonop.get("noi_grant", 0.0),
        "NOE_INT": nonop.get("noe_int", 0.0),
        "NOE_OTH": nonop.get("noe_oth", 0.0),
        "ORD": ord_v,
    }


def _required_sales_for_ord(target_ord: float, plan: dict, nonop: dict) -> float:
    gp = max(1e-9, float(plan["gp_rate"]))
    opex_ttl = plan["opex_fixed"] + plan["opex_h"] + plan["opex_dep"] + plan["opex_oth"]
    noi = nonop.get("noi_misc", 0.0) + nonop.get("noi_grant", 0.0)
    noe = nonop.get("noe_int", 0.0) + nonop.get("noe_oth", 0.0)
    return (target_ord + opex_ttl - noi + noe) / gp


def _be_sales(plan: dict, nonop: dict, *, mode: str = "OP") -> float:
    gp = max(1e-9, float(plan["gp_rate"]))
    opex_ttl = plan["opex_fixed"] + plan["opex_h"] + plan["opex_dep"] + plan["opex_oth"]
    noi = nonop.get("noi_misc", 0.0) + nonop.get("noi_grant", 0.0)
    noe = nonop.get("noe_int", 0.0) + nonop.get("noe_oth", 0.0)
    if mode == "ORD":
        return (opex_ttl - noi + noe) / gp
    return opex_ttl / gp


def build_scenario_dataframe(base_plan: dict, plan: dict,
                             nonop: dict | None = None,
                             target_ord: float = 50_000_000,
                             be_mode: str = "OP") -> pd.DataFrame:
    nonop = NONOP_DEFAULT if nonop is None else nonop

    # ① 目標
    res_target = compute_plan(plan)
    col_target = _line_items(res_target, nonop)

    # ②〜④ 売上スケール
    def col_sales_scale(scale: float):
        p = _plan_with(plan, sales=plan["sales"] * scale)
        return _line_items(compute_plan(p), nonop)

    col_sales_up10 = col_sales_scale(1.10)
    col_sales_dn5 = col_sales_scale(0.95)
    col_sales_dn10 = col_sales_scale(0.90)

    # ⑤ 粗利率+1pt
    p_gp_up = _plan_with(plan, gp_rate=min(1.0, plan["gp_rate"] + 0.01))
    col_gp_up = _line_items(compute_plan(p_gp_up), nonop)

    # ⑥ 目標経常
    req_sales = max(0.0, _required_sales_for_ord(target_ord, plan, nonop))
    p_ord = _plan_with(plan, sales=req_sales)
    col_ord = _line_items(compute_plan(p_ord), nonop)

    # ⑦ 昨年同一
    col_last = _line_items(compute_plan(base_plan), nonop)

    # ⑧ 損益分岐点売上
    be_sales = max(0.0, _be_sales(plan, nonop, mode=be_mode))
    p_be = _plan_with(plan, sales=be_sales)
    col_be = _line_items(compute_plan(p_be), nonop)

    df = pd.DataFrame.from_dict({
        "目標": col_target,
        "売上高10%増": col_sales_up10,
        "売上高5%減": col_sales_dn5,
        "売上高10%減": col_sales_dn10,
        "粗利率+1pt": col_gp_up,
        "経常利益5千万円": col_ord,
        "昨年同一": col_last,
        "損益分岐点売上高": col_be,
    }, orient="index").T
    return df


def render_scenario_table(base_plan: dict, plan: dict,
                          nonop: dict | None = None,
                          *, target_ord: float = 50_000_000,
                          be_mode: str = "OP"):
    st.subheader("📊 シナリオ比較（是正版）")
    df = build_scenario_dataframe(base_plan, plan, nonop, target_ord, be_mode)
    st.dataframe(df.style.format("{:,.0f}"), use_container_width=True)

class PlanConfig:
    def __init__(self, base_sales: float, fte: float, unit: str) -> None:
        self.base_sales = base_sales
        self.fte = max(0.0001, fte)
        self.unit = unit
        self.items: Dict[str, Dict[str, float]] = {}

    def set_rate(self, code: str, rate: float, rate_base: str = 'sales') -> None:
        self.items[code] = {"method": "rate", "value": float(rate), "rate_base": rate_base}

    def set_amount(self, code: str, amount: float) -> None:
        self.items[code] = {"method": "amount", "value": float(amount), "rate_base": "fixed"}

    def clone(self) -> 'PlanConfig':
        c = PlanConfig(self.base_sales, self.fte, self.unit)
        c.items = {k: v.copy() for k, v in self.items.items()}
        return c


def dual_input_row(label: str, base_sales: float, *,
                   mode_key: str,
                   pct_default: float = 0.0,
                   amount_default: float = 0.0,
                   pct_min: float = 0.0, pct_max: float = 3.0, pct_step: float = 0.005,
                   help_text: str = "") -> dict:
    """
    返り値: {"method": "rate" or "amount", "value": float}
    - mode=="％（増減/売上対比）": 率を編集、実額は参考表示（= rate * base_sales）
    - mode=="実額（円）": 実額を編集、率は参考表示（= amount / base_sales）
    - 0除算/NaNは自動で保護し、表示は0とする
    """
    mode = st.session_state.get(mode_key, "％（増減/売上対比）")
    key_base = label.replace("｜", "_").replace(" ", "_")
    if mode == "％（増減/売上対比）":
        rate = st.number_input(
            f"{label}（率）",
            min_value=pct_min,
            max_value=pct_max,
            step=pct_step,
            format="%.3f",
            value=pct_default,
            help=help_text,
            key=f"{key_base}_pct"
        )
        amount = rate * base_sales
        if not math.isfinite(amount):
            amount = 0.0
        st.caption(f"金額 ¥{amount:,.0f}")
        return {"method": "rate", "value": rate}
    else:
        amount = st.number_input(
            f"{label}（実額）",
            min_value=0.0,
            step=1_000_000.0,
            format="%.0f",
            value=amount_default,
            help=help_text,
            key=f"{key_base}_amt"
        )
        if not math.isfinite(amount):
            amount = 0.0
        rate = amount / base_sales if base_sales > 0 else 0.0
        if not math.isfinite(rate):
            rate = 0.0
        st.caption(f"率 {rate*100:.0f}%")
        return {"method": "amount", "value": amount}

def compute(plan: PlanConfig, sales_override: float | None = None, amount_overrides: Dict[str, float] | None = None) -> Dict[str, float]:
    S = float(plan.base_sales if sales_override is None else sales_override)
    amt = {code: 0.0 for code, *_ in ITEMS}
    amt["REV"] = S

    def line_amount(code, gross_guess):
        cfg = plan.items.get(code, None)
        if amount_overrides and code in amount_overrides:
            return float(amount_overrides[code])
        if cfg is None:
            return 0.0
        if cfg["method"] == "amount":
            return float(cfg["value"])
        r = float(cfg["value"])
        base = cfg.get("rate_base", "sales")
        if base == "sales":
            return S * r
        elif base == "gross":
            return max(0.0, gross_guess) * r
        elif base == "fixed":
            return r
        return S * r

    cogs_codes = ["COGS_MAT", "COGS_LBR", "COGS_OUT_SRC", "COGS_OUT_CON", "COGS_OTH"]
    sales_based_cogs = 0.0
    for code in cogs_codes:
        cfg = plan.items.get(code)
        if cfg and cfg["method"] == "rate" and cfg.get("rate_base", "sales") == "sales":
            sales_based_cogs += S * float(cfg["value"])
        elif cfg and cfg["method"] == "amount":
            sales_based_cogs += float(cfg["value"])

    gross = S - sales_based_cogs
    for _ in range(5):
        cogs = 0.0
        for code in cogs_codes:
            cogs += max(0.0, line_amount(code, gross))
        gross_new = S - cogs
        if abs(gross_new - gross) < 1e-6:
            gross = gross_new
            break
        gross = gross_new

    cogs_total = 0.0
    for code in cogs_codes:
        val = max(0.0, line_amount(code, gross))
        amt[code] = val
        cogs_total += val
    amt["COGS_TTL"] = cogs_total
    amt["GROSS"] = S - cogs_total

    opex_codes = ["OPEX_H", "OPEX_K", "OPEX_DEP"]
    opex_total = 0.0
    for code in opex_codes:
        val = max(0.0, line_amount(code, amt["GROSS"]))
        amt[code] = val
        opex_total += val
    amt["OPEX_TTL"] = opex_total

    amt["OP"] = amt["GROSS"] - amt["OPEX_TTL"]

    noi_codes = ["NOI_MISC", "NOI_GRANT", "NOI_OTH"]
    noe_codes = ["NOE_INT", "NOE_OTH"]
    for code in noi_codes + noe_codes:
        val = max(0.0, line_amount(code, amt["GROSS"]))
        amt[code] = val

    amt["ORD"] = amt["OP"] + (amt["NOI_MISC"] + amt["NOI_GRANT"] + amt["NOI_OTH"]) - (amt["NOE_INT"] + amt["NOE_OTH"])

    var_cost = 0.0
    for code in cogs_codes + opex_codes + noi_codes + noe_codes:
        cfg = plan.items.get(code)
        if cfg and cfg["method"] == "rate" and cfg.get("rate_base", "sales") in ("sales", "gross"):
            if cfg.get("rate_base") == "gross":
                g_ratio = amt["GROSS"] / S if S > 0 else 0.0
                var_cost += S * (cfg["value"] * g_ratio)
            else:
                var_cost += S * cfg["value"]

    fixed_cost = 0.0
    for code in cogs_codes + opex_codes + noi_codes + noe_codes:
        cfg = plan.items.get(code)
        if cfg and cfg["method"] == "amount":
            fixed_cost += cfg["value"]
        elif cfg and cfg.get("rate_base") == "fixed":
            if cfg["method"] == "rate":
                fixed_cost += cfg["value"]
    cm_ratio = 1.0 - (var_cost / S if S > 0 else 0.0)
    if cm_ratio <= 0:
        be_sales = float("inf")
    else:
        be_sales = fixed_cost / cm_ratio
    amt["BE_SALES"] = be_sales

    fte = max(0.0001, plan.fte)
    amt["PC_SALES"] = amt["REV"] / fte
    amt["PC_GROSS"] = amt["GROSS"] / fte
    amt["PC_ORD"] = amt["ORD"] / fte
    amt["LDR"] = (amt["OPEX_H"] / amt["GROSS"]) if amt["GROSS"] > 0 else np.nan

    return amt

def bisection_for_target_op(plan: PlanConfig, target_op: float, s_low: float, s_high: float, max_iter=60, eps=1_000.0) -> Tuple[float, Dict[str, float]]:
    def op_at(S):
        return compute(plan, sales_override=S)["ORD"]
    low, high = max(0.0, s_low), max(s_low * 1.5, s_high)
    f_low = op_at(low)
    f_high = op_at(high)
    it = 0
    while (f_low - target_op) * (f_high - target_op) > 0 and high < 1e13 and it < 40:
        high = high * 1.6 if high > 0 else 1_000_000.0
        f_high = op_at(high)
        it += 1
    for _ in range(max_iter):
        mid = 0.5 * (low + high)
        f_mid = op_at(mid)
        if abs(f_mid - target_op) <= eps:
            return mid, compute(plan, sales_override=mid)
        if (f_low - target_op) * (f_mid - target_op) <= 0:
            high, f_high = mid, f_mid
        else:
            low, f_low = mid, f_mid
    mid = 0.5 * (low + high)
    return mid, compute(plan, sales_override=mid)

st.markdown("## 🧭 マネジメント・コントロールハブ")
with st.container(border=True):
    st.caption("率と実額を切り替えながら、重要な経営レバーを中央エリアで一括コントロールできます。")
    base_cols = st.columns([2.4, 1.3, 1.3], gap="large")
    with base_cols[0]:
        mode = st.radio(
            "入力モード",
            ["％（増減/売上対比）", "実額（円）"],
            horizontal=True,
            index=0,
            key="input_mode",
        )
        st.caption("％指定で売上に対する構成比を直感的に管理。必要に応じてワンクリックで実額モードへ。")
    with base_cols[1]:
        fiscal_year = st.number_input("会計年度", value=int(DEFAULTS["fiscal_year"]), step=1, format="%d")
        unit = st.selectbox("表示単位", ["百万円", "千円", "円"], index=0, help="計算は円ベース、表示のみ丸めます。")
    with base_cols[2]:
        base_sales = st.number_input(
            "売上高（ベース）",
            value=float(DEFAULTS["sales"]),
            step=10_000_000.0,
            min_value=0.0,
            format="%.0f",
        )
        fte = st.number_input("人員数（FTE換算）", value=float(DEFAULTS["fte"]), step=1.0, min_value=0.0)

    st.markdown("#### 🎚️ コスト & 収益レバー")
    st.caption("主要コストは3つのタブに整理。カテゴリごとにまとめたカードで、配分バランスを素早く再設計できます。")
    tab_cost, tab_internal, tab_nonop = st.tabs(["外部仕入", "内部費用", "営業外 / 営業外費用"])

    with tab_cost:
        ext_row1 = st.columns(3, gap="large")
        with ext_row1[0]:
            cogs_mat_input = dual_input_row(
                "材料費",
                base_sales,
                mode_key="input_mode",
                pct_default=float(DEFAULTS["cogs_mat_rate"]),
                amount_default=base_sales * DEFAULTS["cogs_mat_rate"],
                pct_step=0.01,
            )
        with ext_row1[1]:
            cogs_lbr_input = dual_input_row(
                "労務費(外部)",
                base_sales,
                mode_key="input_mode",
                pct_default=float(DEFAULTS["cogs_lbr_rate"]),
                amount_default=base_sales * DEFAULTS["cogs_lbr_rate"],
                pct_step=0.01,
            )
        with ext_row1[2]:
            cogs_out_src_input = dual_input_row(
                "外注費(専属)",
                base_sales,
                mode_key="input_mode",
                pct_default=float(DEFAULTS["cogs_out_src_rate"]),
                amount_default=base_sales * DEFAULTS["cogs_out_src_rate"],
                pct_step=0.01,
            )
        ext_row2 = st.columns(2, gap="large")
        with ext_row2[0]:
            cogs_out_con_input = dual_input_row(
                "外注費(委託)",
                base_sales,
                mode_key="input_mode",
                pct_default=float(DEFAULTS["cogs_out_con_rate"]),
                amount_default=base_sales * DEFAULTS["cogs_out_con_rate"],
                pct_step=0.01,
            )
        with ext_row2[1]:
            cogs_oth_input = dual_input_row(
                "その他諸経費",
                base_sales,
                mode_key="input_mode",
                pct_default=float(DEFAULTS["cogs_oth_rate"]),
                amount_default=base_sales * DEFAULTS["cogs_oth_rate"],
                pct_step=0.005,
            )

    with tab_internal:
        int_row = st.columns(3, gap="large")
        with int_row[0]:
            opex_h_input = dual_input_row(
                "人件費",
                base_sales,
                mode_key="input_mode",
                pct_default=float(DEFAULTS["opex_h_rate"]),
                amount_default=base_sales * DEFAULTS["opex_h_rate"],
                pct_step=0.01,
            )
        with int_row[1]:
            opex_k_input = dual_input_row(
                "経費",
                base_sales,
                mode_key="input_mode",
                pct_default=float(DEFAULTS["opex_k_rate"]),
                amount_default=base_sales * DEFAULTS["opex_k_rate"],
                pct_step=0.01,
            )
        with int_row[2]:
            opex_dep_input = dual_input_row(
                "減価償却",
                base_sales,
                mode_key="input_mode",
                pct_default=float(DEFAULTS["opex_dep_rate"]),
                amount_default=base_sales * DEFAULTS["opex_dep_rate"],
                pct_step=0.001,
            )

    with tab_nonop:
        nonop_row1 = st.columns(3, gap="large")
        with nonop_row1[0]:
            noi_misc_input = dual_input_row(
                "営業外収益：雑収入",
                base_sales,
                mode_key="input_mode",
                pct_default=float(DEFAULTS["noi_misc_rate"]),
                amount_default=base_sales * DEFAULTS["noi_misc_rate"],
                pct_min=0.0,
                pct_max=1.0,
                pct_step=0.0005,
            )
        with nonop_row1[1]:
            noi_grant_input = dual_input_row(
                "営業外収益：補助金",
                base_sales,
                mode_key="input_mode",
                pct_default=float(DEFAULTS["noi_grant_rate"]),
                amount_default=base_sales * DEFAULTS["noi_grant_rate"],
                pct_min=0.0,
                pct_max=1.0,
                pct_step=0.0005,
            )
        with nonop_row1[2]:
            noi_oth_input = dual_input_row(
                "営業外収益：その他",
                base_sales,
                mode_key="input_mode",
                pct_default=float(DEFAULTS["noi_oth_rate"]),
                amount_default=base_sales * DEFAULTS["noi_oth_rate"],
                pct_min=0.0,
                pct_max=1.0,
                pct_step=0.0005,
            )
        nonop_row2 = st.columns(2, gap="large")
        with nonop_row2[0]:
            noe_int_input = dual_input_row(
                "営業外費用：支払利息",
                base_sales,
                mode_key="input_mode",
                pct_default=float(DEFAULTS["noe_int_rate"]),
                amount_default=base_sales * DEFAULTS["noe_int_rate"],
                pct_min=0.0,
                pct_max=1.0,
                pct_step=0.0005,
            )
        with nonop_row2[1]:
            noe_oth_input = dual_input_row(
                "営業外費用：雑損",
                base_sales,
                mode_key="input_mode",
                pct_default=float(DEFAULTS["noe_oth_rate"]),
                amount_default=base_sales * DEFAULTS["noe_oth_rate"],
                pct_min=0.0,
                pct_max=1.0,
                pct_step=0.0005,
            )

with st.expander("🎨 グラフスタイル", expanded=False):
    st.caption("トルネード図やウォーターフォールなどのビジュアルテーマを、ブランドカラーに合わせて細かく調整できます。")
    style_cols = st.columns(3, gap="large")
    with style_cols[0]:
        fig_bg = st.color_picker("図背景色", PLOT_STYLE_DEFAULT["figure_bg"])
        axes_bg = st.color_picker("枠背景色", PLOT_STYLE_DEFAULT["axes_bg"])
        show_grid = st.checkbox("グリッド線を表示", value=PLOT_STYLE_DEFAULT["grid"])
    with style_cols[1]:
        grid_color = st.color_picker("グリッド線色", PLOT_STYLE_DEFAULT["grid_color"])
        pos_color = st.color_picker("増加色", PLOT_STYLE_DEFAULT["pos_color"])
        neg_color = st.color_picker("減少色", PLOT_STYLE_DEFAULT["neg_color"])
    with style_cols[2]:
        node_size = st.slider("ノードサイズ", 1, 30, PLOT_STYLE_DEFAULT["node_size"])
        font_color = st.color_picker("フォント色", PLOT_STYLE_DEFAULT["font_color"])
        font_size = st.slider("フォントサイズ", 6, 24, PLOT_STYLE_DEFAULT["font_size"])
        alpha = st.slider("透過度", 0.0, 1.0, PLOT_STYLE_DEFAULT["alpha"], 0.05)



plot_style = {
    "figure_bg": fig_bg,
    "axes_bg": axes_bg,
    "grid": show_grid,
    "grid_color": grid_color,
    "pos_color": pos_color,
    "neg_color": neg_color,
    "node_size": node_size,
    "font_color": font_color,
    "font_size": font_size,
    "alpha": alpha,
}

base_plan = PlanConfig(base_sales=base_sales, fte=fte, unit=unit)


def apply_setting(code: str, result: dict) -> None:
    if result["method"] == "rate":
        base_plan.set_rate(code, result["value"], "sales")
    else:
        base_plan.set_amount(code, result["value"])


apply_setting("COGS_MAT", cogs_mat_input)
apply_setting("COGS_LBR", cogs_lbr_input)
apply_setting("COGS_OUT_SRC", cogs_out_src_input)
apply_setting("COGS_OUT_CON", cogs_out_con_input)
apply_setting("COGS_OTH", cogs_oth_input)

apply_setting("OPEX_H", opex_h_input)
apply_setting("OPEX_K", opex_k_input)
apply_setting("OPEX_DEP", opex_dep_input)

apply_setting("NOI_MISC", noi_misc_input)
apply_setting("NOI_GRANT", noi_grant_input)
apply_setting("NOI_OTH", noi_oth_input)
apply_setting("NOE_INT", noe_int_input)
apply_setting("NOE_OTH", noe_oth_input)

tab_input, tab_scen, tab_analysis, tab_ai, tab_export = st.tabs(
    ["📝 計画入力", "🧪 シナリオ", "📊 感応度分析", "🤖 AIインサイト", "📤 エクスポート"]
)

with tab_input:
    st.subheader("単年利益計画（目標列）")
    base_amt = compute(base_plan)

    def fmt_amount_with_unit(value: float) -> str:
        formatted = format_money(value, base_plan.unit)
        return formatted if formatted == "—" else f"{formatted} {base_plan.unit}"

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("売上高", fmt_amount_with_unit(base_amt["REV"]))
    c2.metric("粗利(加工高)", fmt_amount_with_unit(base_amt["GROSS"]))
    c3.metric("営業利益", fmt_amount_with_unit(base_amt["OP"]))
    c4.metric("経常利益", fmt_amount_with_unit(base_amt["ORD"]))
    be_label = "∞" if not math.isfinite(base_amt["BE_SALES"]) else fmt_amount_with_unit(base_amt["BE_SALES"])
    c5.metric("損益分岐点売上高", be_label)

    c6, c7, c8 = st.columns(3)
    c6.metric("一人当たり売上", fmt_amount_with_unit(base_amt["PC_SALES"]))
    c7.metric("一人当たり粗利", fmt_amount_with_unit(base_amt["PC_GROSS"]))
    ldr = base_amt["LDR"]
    ldr_str = "—" if (ldr is None or not math.isfinite(ldr)) else f"{ldr*100:.0f}%"
    c8.metric("労働分配率", ldr_str)

    st.markdown("### 標準原価の見える化（中央ビュー）")
    st.caption("コントロールハブで設定した原価や費用がリアルタイムに反映され、売上に対するインパクトを一目で確認できます。")

    revenue = float(base_amt.get("REV", 0.0))
    cost_cards = []
    for code, label, desc, extra_class in COST_PILL_ITEMS:
        value = float(base_amt.get(code, 0.0) or 0.0)
        ratio = value / revenue if revenue else float("nan")
        cost_cards.append({
            "code": code,
            "label": label,
            "desc": desc,
            "value": value,
            "ratio": ratio,
            "class": extra_class,
        })

    pill_columns = st.columns(3)
    for idx, card in enumerate(cost_cards):
        col = pill_columns[idx % 3]
        ratio_text = format_ratio(card["ratio"])
        amount_text = fmt_amount_with_unit(card["value"])
        pill_class = "cost-pill"
        if card["class"]:
            pill_class = f"{pill_class} {card['class']}"
        pill_html = (
            f"<div class='{pill_class}'>"
            f"<strong>{card['label']}</strong>"
            f"<span>{amount_text}</span>"
            f"<small>{ratio_text} ／ {card['desc']}</small>"
            "</div>"
        )
        col.markdown(pill_html, unsafe_allow_html=True)

    cost_chart_cards = [
        card
        for card in cost_cards
        if card["code"] in {"COGS_MAT", "COGS_LBR", "COGS_OUT_SRC", "COGS_OUT_CON", "COGS_OTH"}
    ]
    if revenue > 0 and any(card["value"] > 0 for card in cost_chart_cards):
        names = [card["label"] for card in cost_chart_cards]
        shares = [max(0.0, card["ratio"]) * 100 if math.isfinite(card["ratio"]) else 0.0 for card in cost_chart_cards]
        max_share = max(shares) if shares else 0.0
        slider_min = 5.0
        slider_max = max(
            slider_min + 5.0,
            (math.ceil(max_share * 1.6 / 5.0) * 5.0) if max_share > 0 else 30.0,
        )
        default_limit = max(
            slider_min + 5.0,
            (math.ceil(max_share * 1.2 / 5.0) * 5.0) if max_share > 0 else 25.0,
        )
        share_axis_max = st.slider(
            "表示上限（%）",
            min_value=float(slider_min),
            max_value=float(slider_max),
            value=float(min(default_limit, slider_max)),
            step=1.0,
            key="cost_share_axis",
            help="棒グラフ右端のスケールをコントロールできます。",
        )

        colors = [THEME_COLORS["primary_light"] if i % 2 == 0 else THEME_COLORS["primary"] for i in range(len(names))]
        hover_details = [
            f"{format_ratio(card['ratio'])} ／ {fmt_amount_with_unit(card['value'])}"
            for card in cost_chart_cards
        ]
        fig_height = 120 + 70 * len(cost_chart_cards)
        fig = go.Figure(
            data=[
                go.Bar(
                    x=shares,
                    y=names,
                    orientation="h",
                    marker=dict(
                        color=colors,
                        line=dict(color="rgba(31, 78, 121, 0.18)", width=1.4),
                    ),
                    text=[format_ratio(card["ratio"]) for card in cost_chart_cards],
                    textposition="outside",
                    textfont=dict(size=12, color=THEME_COLORS["text"]),
                    customdata=hover_details,
                    hovertemplate="<b>%{y}</b><br>売上比率: %{x:.1f}%<br>%{customdata}<extra></extra>",
                    cliponaxis=False,
                )
            ]
        )
        fig.update_layout(
            height=fig_height,
            margin=dict(l=0, r=18, t=48, b=10),
            bargap=0.25,
            plot_bgcolor="#FFFFFF",
            paper_bgcolor="#FFFFFF",
            xaxis=dict(
                title="売上比率（%）",
                range=[0, share_axis_max],
                showgrid=True,
                gridcolor="#D4DEE9",
                ticksuffix="%",
                zeroline=False,
                rangeslider=dict(visible=True, thickness=0.12, bgcolor="rgba(31, 78, 121, 0.08)"),
            ),
            yaxis=dict(autorange="reversed", showgrid=False),
            hoverlabel=dict(bgcolor=THEME_COLORS["primary"], font=dict(color="#FFFFFF")),
        )
        st.plotly_chart(
            fig,
            use_container_width=True,
            config={
                "displaylogo": False,
                "modeBarButtonsToAdd": ["drawline", "drawrect", "eraseshape"],
                "toImageButtonOptions": {"filename": "standard-cost-breakdown"},
            },
        )
        st.caption(
            "横棒グラフは売上100に対し、それぞれの標準原価がどれだけを占めるかを示します。ズーム/パンに加え、スライダーで目盛りをコントロールできます。"
        )

    cost_table = [
        {
            "コスト項目": card["label"],
            "売上比率": format_ratio(card["ratio"]),
            "金額": fmt_amount_with_unit(card["value"]),
            "ひとことで": card["desc"],
        }
        for card in cost_cards
    ]
    st.dataframe(
        pd.DataFrame(cost_table),
        use_container_width=True,
        hide_index=True,
    )
    st.caption("カードと表はコントロールハブの入力に連動して更新されます。粗利（CT）と標準原価のバランスを中央ビューで確認してください。")

    st.markdown("### 主要項目（経営メモ付き）")
    rows = []
    for code, label, group in ITEMS:
        if code in ("PC_SALES", "PC_GROSS", "PC_ORD", "LDR", "BE_SALES"):
            continue
        val = base_amt[code]
        memo = PLAIN_LANGUAGE.get(code, "—")
        rows.append({
            "項目": label,
            "経営メモ": memo,
            "金額": fmt_amount_with_unit(val),
        })
    df = pd.DataFrame(rows)
    st.dataframe(
        df,
        use_container_width=True,
        hide_index=True,
        height=min(520, 40 + 28 * len(rows)),
    )

    st.info(
        "ヒント: コントロールハブの％／実額・人員・売上を調整すると、標準原価ビューと一覧表が即座に更新されます。固定費や個別額を設定したい場合は、下の『金額上書き』をご利用ください。"
    )

    with st.expander("🔧 金額上書き（固定費/個別額の設定）", expanded=False):
        st.caption("金額が入力された項目は、率の指定より優先され固定費扱いになります。")
        col1, col2, col3 = st.columns(3)
        override_inputs = {}
        for i, code in enumerate(["COGS_MAT","COGS_LBR","COGS_OUT_SRC","COGS_OUT_CON","COGS_OTH","OPEX_H","OPEX_K","OPEX_DEP","NOI_MISC","NOI_GRANT","NOI_OTH","NOE_INT","NOE_OTH"]):
            if i % 3 == 0:
                c = col1
            elif i % 3 == 1:
                c = col2
            else:
                c = col3
            # Look up label without reconstructing the dictionary each time
            val = c.number_input(
                f"{ITEM_LABELS[code]}（金額上書き）",
                min_value=0.0,
                value=0.0,
                step=1_000_000.0,
                key=f"ov_{code}"
            )
            if val > 0:
                override_inputs[code] = val

        if st.button("上書きを反映", type="primary"):
            preview_amt = compute(base_plan, amount_overrides=override_inputs)
            st.session_state["overrides"] = override_inputs
            st.success("上書きを反映しました（この状態でシナリオにも適用されます）。")

            rows2 = []
            for code, label, group in ITEMS:
                if code in ("PC_SALES","PC_GROSS","PC_ORD","LDR","BE_SALES"):
                    continue
                before = base_amt[code]
                after = preview_amt[code]
                rows2.append({
                    "項目": label,
                    "経営メモ": PLAIN_LANGUAGE.get(code, "—"),
                    "前": fmt_amount_with_unit(before),
                    "後": fmt_amount_with_unit(after),
                })
            st.dataframe(pd.DataFrame(rows2), use_container_width=True, hide_index=True)

    glossary_html = "<div class='glossary-card'><h4>用語ミニガイド</h4><ul>"
    for item in GLOSSARY_ITEMS:
        glossary_html += f"<li><strong>{item['term']}</strong><span>{item['description']}</span></li>"
    glossary_html += "</ul></div>"
    st.markdown(glossary_html, unsafe_allow_html=True)

def scenario_table(plan: PlanConfig, unit: str, overrides: Dict[str, float]) -> Tuple[pd.DataFrame, pd.DataFrame, List[Tuple[str, Dict[str, float]]]]:
    # --- SCENARIO UX
    type_display = ["なし", "売上高±%", "粗利率±pt", "目標経常", "昨年同一", "BEP"]
    type_map = {"なし": "none", "売上高±%": "sales_pct", "粗利率±pt": "gross_pt", "目標経常": "target_op", "昨年同一": "last_year", "BEP": "bep"}
    default_specs = [
        {"名称": "目標", "タイプ": "なし", "値": None},
        {"名称": "売上高10%増", "タイプ": "売上高±%", "値": 10.0},
        {"名称": "売上高5%減", "タイプ": "売上高±%", "値": -5.0},
        {"名称": "売上高10%減", "タイプ": "売上高±%", "値": -10.0},
        {"名称": "粗利1%減", "タイプ": "粗利率±pt", "値": -1.0},
        {"名称": "経常利益5千万円", "タイプ": "目標経常", "値": 50_000_000.0},
        {"名称": "昨年同一", "タイプ": "昨年同一", "値": None},
        {"名称": "損益分岐点売上高", "タイプ": "BEP", "値": None},
    ]
    df = st.session_state.get("scenario_df")
    if df is None:
        df = pd.DataFrame(default_specs)
    st.caption("各シナリオのラベルとパラメータを編集できます。")
    editor = st.data_editor(
        df,
        key="scenario_editor",
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            "名称": st.column_config.TextColumn("名称"),
            "タイプ": st.column_config.SelectboxColumn("タイプ", options=type_display),
            "値": st.column_config.NumberColumn("値", help="タイプにより入力範囲が異なります"),
        },
    )
    st.session_state["scenario_df"] = editor.copy()

    def fmt_with_unit(value: float) -> str:
        text = format_money(value, unit)
        return text if text == "—" else f"{text} {unit}"

    def apply_driver(plan: PlanConfig, spec: Dict[str, float], overrides_local: Dict[str, float]):
        t = spec["type"]
        v = spec.get("value", None)
        if t == "none":
            return plan.base_sales, overrides_local, None
        if t == "sales_pct":
            S = plan.base_sales * (1.0 + float(v))
            return S, overrides_local, None
        if t == "gross_pt":
            delta = float(v)
            S = plan.base_sales
            delta_e = -delta * S
            ov = dict(overrides_local) if overrides_local else {}
            current = ov.get("COGS_OTH", None)
            if current is None:
                tmp = compute(plan, sales_override=S, amount_overrides=ov)
                base_oth = tmp["COGS_OTH"]
                ov["COGS_OTH"] = max(0.0, base_oth + delta_e)
            else:
                ov["COGS_OTH"] = max(0.0, current + delta_e)
            return S, ov, None
        if t == "target_op":
            target = float(v)
            sol_S, sol_amt = bisection_for_target_op(plan, target, s_low=0.0, s_high=max(1.2 * plan.base_sales, 1_000_000.0))
            return sol_S, overrides_local, sol_amt
        if t == "last_year":
            return plan.base_sales, overrides_local, None
        if t == "bep":
            temp = compute(plan, sales_override=plan.base_sales, amount_overrides=overrides_local)
            be = temp["BE_SALES"]
            return be if math.isfinite(be) else plan.base_sales, overrides_local, None
        return plan.base_sales, overrides_local, None

    b1, b2, b3, b4, b5 = st.columns(5)
    if b1.button("➕ 追加"):
        new_name = f"シナリオ{len(editor)+1}"
        editor.loc[len(editor)] = [new_name, "なし", None]
        st.session_state["scenario_df"] = editor
    if b2.button("🗑️ 選択行を削除"):
        sel = st.session_state.get("scenario_editor", {}).get("selected_rows", [])
        if sel:
            editor = editor.drop(index=sel).reset_index(drop=True)
            st.session_state["scenario_df"] = editor
    if b3.button("⟳ 既定にリセット"):
        editor = pd.DataFrame(default_specs)
        st.session_state["scenario_df"] = editor
    if b4.button("📌 保存"):
        st.session_state["scenarios"] = editor.to_dict(orient="records")
        st.success("保存しました。")
    if b5.button("📥 読込") and "scenarios" in st.session_state:
        editor = pd.DataFrame(st.session_state["scenarios"])
        st.session_state["scenario_df"] = editor

    selected = st.session_state.get("scenario_editor", {}).get("selected_rows", [])
    if len(selected) == 1:
        idx = selected[0]
        row = editor.loc[idx]
        typ_code = type_map.get(row["タイプ"], "none")
        with st.expander(f"詳細設定：{row['名称']}", expanded=True):
            if typ_code == "sales_pct":
                val = st.slider("売上高±%", -50.0, 50.0, float(row["値"] or 0.0), 1.0)
                editor.at[idx, "値"] = val
            elif typ_code == "gross_pt":
                val = st.slider("粗利率±pt", -10.0, 10.0, float(row["値"] or 0.0), 0.5, help="1pt=1%ポイント")
                editor.at[idx, "値"] = val
            elif typ_code == "target_op":
                val = st.number_input("目標経常利益（円）", min_value=0.0, value=float(row["値"] or 0.0), step=1_000_000.0, format="%.0f")
                editor.at[idx, "値"] = val
            else:
                st.write("—")
        st.session_state["scenario_df"] = editor
        spec = {"type": typ_code, "value": editor.at[idx, "値"]}
        base_amt = compute(plan, amount_overrides=overrides)
        S_override, ov, pre_amt = apply_driver(plan, spec, overrides)
        amt_prev = compute(plan, sales_override=S_override, amount_overrides=ov) if pre_amt is None else pre_amt
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("売上高", fmt_with_unit(amt_prev["REV"]))
        c2.metric("粗利（CT）", fmt_with_unit(amt_prev["GROSS"]))
        c3.metric("経常利益", fmt_with_unit(amt_prev["ORD"]))
        be_lbl = "∞" if not math.isfinite(amt_prev["BE_SALES"]) else fmt_with_unit(amt_prev["BE_SALES"])
        c4.metric("損益分岐点売上高", be_lbl)

    editable = []
    for _, row in editor.iterrows():
        typ_code = type_map.get(row["タイプ"], "none")
        val = row["値"]
        val = None if val is None or (isinstance(val, float) and (np.isnan(val) or np.isinf(val))) else float(val)
        editable.append((row["名称"], {"type": typ_code, "value": val}))

    cols = ["項目", "経営メモ"] + [nm for nm, _ in editable]
    rows = {
        code: [label, PLAIN_LANGUAGE.get(code, "—")]
        for code, label, _ in ITEMS
        if code not in ("PC_SALES", "PC_GROSS", "PC_ORD", "LDR", "BE_SALES")
    }
    kpis = {
        "BE_SALES": ["損益分岐点売上高", PLAIN_LANGUAGE.get("BE_SALES", "—")],
        "PC_SALES": ["一人当たり売上", PLAIN_LANGUAGE.get("PC_SALES", "—")],
        "PC_GROSS": ["一人当たり粗利", PLAIN_LANGUAGE.get("PC_GROSS", "—")],
        "PC_ORD": ["一人当たり経常利益", PLAIN_LANGUAGE.get("PC_ORD", "—")],
        "LDR": ["労働分配率", PLAIN_LANGUAGE.get("LDR", "—")],
    }

    base_amt = compute(plan, amount_overrides=overrides)
    for code, label, _ in ITEMS:
        if code in rows:
            rows[code].append(format_money(base_amt.get(code, 0.0), unit))
    for k in kpis.keys():
        if k == "LDR":
            val = base_amt.get("LDR", float("nan"))
            kpis[k].append(f"{val*100:.0f}%" if val == val else "—")
        else:
            kpis[k].append(format_money(base_amt.get(k, 0.0), unit))

    for nm, spec in editable[1:]:
        S_override, ov, pre_amt = apply_driver(plan, spec, overrides)
        scn_amt = compute(plan, sales_override=S_override, amount_overrides=ov) if pre_amt is None else pre_amt
        for code, label, _ in ITEMS:
            if code in rows:
                rows[code].append(format_money(scn_amt.get(code, 0.0), unit))
        for k in kpis.keys():
            if k == "LDR":
                v = scn_amt.get("LDR", float("nan"))
                kpis[k].append(f"{v*100:.0f}%" if v == v else "—")
            else:
                kpis[k].append(format_money(scn_amt.get(k, 0.0), unit))

    df1 = pd.DataFrame(rows.values(), columns=cols, index=rows.keys())
    df2 = pd.DataFrame(kpis.values(), columns=cols, index=kpis.keys())
    st.subheader("シナリオ比較（金額）")
    st.dataframe(df1, use_container_width=True, hide_index=True)
    st.subheader("KPI（損益分岐点・一人当たり・労働分配率）")
    st.dataframe(df2, use_container_width=True, hide_index=True)
    return df1, df2, editable


def compute_scenario_numeric(plan: PlanConfig, specs: List[Tuple[str, Dict[str, float]]], overrides: Dict[str, float]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """シナリオ比較の数値版（Excel出力やAI分析で再利用）。"""

    cols = ["項目"] + [nm for nm, _ in specs]
    num_rows = {code: [label] for code, label, _ in ITEMS if code not in ("PC_SALES", "PC_GROSS", "PC_ORD", "LDR", "BE_SALES")}
    num_kpis = {
        "BE_SALES": ["損益分岐点売上高"],
        "PC_SALES": ["一人当たり売上"],
        "PC_GROSS": ["一人当たり粗利"],
        "PC_ORD": ["一人当たり経常利益"],
        "LDR": ["労働分配率"],
    }

    def apply_driver(spec: Dict[str, float]):
        t = spec.get("type", "none")
        v = spec.get("value", None)
        if t == "none":
            return plan.base_sales, overrides, None
        if t == "sales_pct":
            return plan.base_sales * (1.0 + float(v)), overrides, None
        if t == "gross_pt":
            S = plan.base_sales
            delta_e = -float(v) * S
            ov = dict(overrides) if overrides else {}
            tmp = compute(plan, sales_override=S, amount_overrides=ov)
            base_oth = tmp["COGS_OTH"]
            ov["COGS_OTH"] = max(0.0, base_oth + delta_e)
            return S, ov, None
        if t == "target_op":
            target = float(v)
            sol_S, sol_amt = bisection_for_target_op(plan, target, s_low=0.0, s_high=max(1.2 * plan.base_sales, 1_000_000.0))
            return sol_S, overrides, sol_amt
        if t == "last_year":
            return plan.base_sales, overrides, None
        if t == "bep":
            temp = compute(plan, sales_override=plan.base_sales, amount_overrides=overrides)
            be = temp["BE_SALES"]
            return (be if math.isfinite(be) else plan.base_sales), overrides, None
        return plan.base_sales, overrides, None

    base_amt = compute(plan, amount_overrides=overrides)
    for code, label, _ in ITEMS:
        if code in num_rows:
            num_rows[code].append(base_amt.get(code, 0.0))
    for k in num_kpis.keys():
        num_kpis[k].append(base_amt.get(k, 0.0))

    for nm, spec in specs[1:]:
        S, ov, pre = apply_driver(spec)
        scn_amt = compute(plan, sales_override=S, amount_overrides=ov) if pre is None else pre
        for code, label, _ in ITEMS:
            if code in num_rows:
                num_rows[code].append(scn_amt.get(code, 0.0))
        for k in num_kpis.keys():
            num_kpis[k].append(scn_amt.get(k, 0.0))

    df_num = pd.DataFrame(num_rows.values(), columns=cols, index=num_rows.keys())
    df_kpi = pd.DataFrame(num_kpis.values(), columns=cols, index=num_kpis.keys())
    return df_num, df_kpi

with tab_scen:
    overrides = st.session_state.get("overrides", {})
    df_amounts, df_kpis, scenario_specs = scenario_table(base_plan, unit, overrides)

numeric_amounts_data, numeric_kpis_data = compute_scenario_numeric(
    base_plan,
    scenario_specs,
    st.session_state.get("overrides", {}),
)

with tab_analysis:
    _set_jp_font()
    base_amt_raw = compute(base_plan)
    base_plan_inputs = {
        "sales": base_amt_raw["REV"],
        "gp_rate": (base_amt_raw["GROSS"] / base_amt_raw["REV"]) if base_amt_raw["REV"] else 0.0,
        "opex_h": base_amt_raw["OPEX_H"],
        "opex_fixed": base_amt_raw["OPEX_K"],
        "opex_dep": base_amt_raw["OPEX_DEP"],
        "opex_oth": -(base_amt_raw["NOI_MISC"] + base_amt_raw["NOI_GRANT"] + base_amt_raw["NOI_OTH"]
                       - base_amt_raw["NOE_INT"] - base_amt_raw["NOE_OTH"]),
    }
    base_amt = compute(base_plan, amount_overrides=st.session_state.get("overrides", {}))
    plan_inputs = {
        "sales": base_amt["REV"],
        "gp_rate": (base_amt["GROSS"] / base_amt["REV"]) if base_amt["REV"] else 0.0,
        "opex_h": base_amt["OPEX_H"],
        "opex_fixed": base_amt["OPEX_K"],
        "opex_dep": base_amt["OPEX_DEP"],
        "opex_oth": -(base_amt["NOI_MISC"] + base_amt["NOI_GRANT"] + base_amt["NOI_OTH"]
                       - base_amt["NOE_INT"] - base_amt["NOE_OTH"]),
    }
    render_scenario_table(base_plan_inputs, plan_inputs, NONOP_DEFAULT,
                          target_ord=50_000_000, be_mode="OP")
    render_sensitivity_view(plan_inputs)

with tab_ai:
    st.markdown("<span class='ai-badge'>AIによる自動レビュー</span>", unsafe_allow_html=True)
    st.subheader("インテリジェント・サマリー")
    st.caption("シナリオやコントロールハブの設定を更新すると、AIインサイトも即座にリフレッシュされます。")
    overrides = st.session_state.get("overrides", {})
    base_amt_ai = compute(base_plan, amount_overrides=overrides)
    metrics = summarize_plan_metrics(base_amt_ai)

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("売上高 (目標)", f"{format_money(metrics['sales'], unit)} {unit}")
    m2.metric("粗利率", format_ratio(metrics.get("gross_margin")))
    m3.metric("経常利益", f"{format_money(metrics['ord'], unit)} {unit}")
    m4.metric("経常利益率", format_ratio(metrics.get("ord_margin")))

    st.markdown("### AIレコメンド")
    insights = generate_ai_recommendations(metrics, numeric_amounts_data, numeric_kpis_data, unit)
    for ins in insights:
        st.markdown(
            f"<div class='insight-card {ins['tone']}'><h4>{ins['title']}</h4><p>{ins['body']}</p></div>",
            unsafe_allow_html=True,
        )

    st.markdown("### 異常値検知 (AI Quality Check)")
    anomalies_df = detect_anomalies_in_plan(numeric_amounts_data, numeric_kpis_data, unit, metrics)
    if not anomalies_df.empty:
        st.dataframe(
            anomalies_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "コメント": st.column_config.TextColumn("コメント", width="large"),
            },
        )
    else:
        st.success("異常値は検出されませんでした。データ整合性は良好です。")

with tab_export:
    st.subheader("エクスポート")
    st.caption("ワンクリックでExcel出力（シート: 金額, KPI, 感応度）。PDFはExcelから印刷設定で作成してください。")
    specs = scenario_specs
    df_num, df_kpi = numeric_amounts_data, numeric_kpis_data

    def recompute_sensitivity_table():
        base_amt = compute(base_plan, amount_overrides=st.session_state.get("overrides", {}))
        base_op = base_amt["ORD"]
        def op_with(ds=0.1, dgp=0.01, dH=0.1, dK=0.1):
            plan = base_plan.clone()
            S = plan.base_sales * (1.0 + ds)
            overrides = st.session_state.get("overrides", {}).copy()
            delta_e = -dgp * S
            overrides["COGS_OTH"] = max(0.0, compute(plan, sales_override=S, amount_overrides=overrides)["COGS_OTH"] + delta_e)
            val = compute(plan, sales_override=S, amount_overrides=overrides)["OPEX_H"]
            overrides["OPEX_H"] = max(0.0, val * (1.0 + dH))
            val = compute(plan, sales_override=S, amount_overrides=overrides)["OPEX_K"]
            overrides["OPEX_K"] = max(0.0, val * (1.0 + dK))
            return compute(plan, sales_override=S, amount_overrides=overrides)["ORD"]
        changes = [
            ("売上高 +10%", op_with(ds=+0.10) - base_op),
            ("売上高 -10%", op_with(ds=-0.10) - base_op),
            ("粗利率 +1pt", op_with(dgp=+0.01) - base_op),
            ("粗利率 -1pt", op_with(dgp=-0.01) - base_op),
            ("人件費 +10%", op_with(dH=+0.10) - base_op),
            ("人件費 -10%", op_with(dH=-0.10) - base_op),
            ("経費 +10%", op_with(dK=+0.10) - base_op),
            ("経費 -10%", op_with(dK=-0.10) - base_op),
        ]
        df = pd.DataFrame(changes, columns=["ドライバ","OP変化（円）"])
        return df

    df_sens = recompute_sensitivity_table()

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheets_written = 0
        if isinstance(df_num, pd.DataFrame) and not df_num.empty:
            df_num.to_excel(writer, sheet_name="金額", index=True)
            sheets_written += 1
        if isinstance(df_kpi, pd.DataFrame) and not df_kpi.empty:
            df_kpi.to_excel(writer, sheet_name="KPI", index=True)
            sheets_written += 1
        if isinstance(df_sens, pd.DataFrame) and not df_sens.empty:
            df_sens.to_excel(writer, sheet_name="感応度", index=False)
            sheets_written += 1
        if sheets_written == 0:
            pd.DataFrame().to_excel(writer, sheet_name="Sheet1")

        wb = writer.book
        if "金額" in wb.sheetnames:
            ws = wb["金額"]
            format_money_and_percent(ws, list(range(2, ws.max_column + 1)), [])
        if "KPI" in wb.sheetnames:
            ws = wb["KPI"]
            money_fmt = "\"¥\"#,##0;[Red]-\"¥\"#,##0"
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=1).value == "労働分配率":
                    for c in range(2, ws.max_column + 1):
                        ws.cell(row=r, column=c).number_format = "0%"
                else:
                    for c in range(2, ws.max_column + 1):
                        ws.cell(row=r, column=c).number_format = money_fmt
        if "感応度" in wb.sheetnames:
            ws = wb["感応度"]
            format_money_and_percent(ws, [2], [])

        meta_ws = wb.create_sheet("メタ情報")
        meta_data = [
            ("作成日時", dt.datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("会計年度", fiscal_year),
            ("表示単位", unit),
            ("FTE", fte),
            ("ベース売上（円）", base_sales),
        ]
        for i, (k, v) in enumerate(meta_data, start=1):
            meta_ws.cell(row=i, column=1, value=k)
            meta_ws.cell(row=i, column=2, value=v)
        format_money_and_percent(meta_ws, [2], [])

        apply_japanese_styles(wb)
    data = output.getvalue()

    st.download_button(
        label="📥 Excel（.xlsx）をダウンロード",
        data=data,
        file_name=f"利益計画_{dt.date.today().isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

st.caption("© 経営計画策定WEBアプリ（Streamlit版） | 表示単位と計算単位を分離し、丸めの影響を最小化しています。")
